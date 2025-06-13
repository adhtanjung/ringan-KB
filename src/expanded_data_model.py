import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_expanded_mental_health_kb_excel():
    """Create an expanded Excel file with comprehensive data for the mental health knowledge base."""
    # Create a new workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Expanded sample data for each sheet
    problems_data = [
        ["problem_id", "problem_name", "description"],
        ["P001", "Anxiety", "Persistent feelings of worry, fear, or nervousness that interfere with daily activities"],
        ["P002", "Depression", "Persistent feelings of sadness, hopelessness, and loss of interest in activities"],
        ["P003", "Stress", "Emotional or physical tension resulting from demanding circumstances"],
        ["P004", "Sleep Issues", "Difficulty falling asleep, staying asleep, or getting restful sleep"],
        ["P005", "Social Isolation", "Lack of social connections and feelings of loneliness"],
        ["P006", "Panic Disorder", "Recurrent unexpected panic attacks and persistent concern about future attacks"],
        ["P007", "PTSD", "Post-traumatic stress disorder caused by experiencing or witnessing a traumatic event"],
        ["P008", "OCD", "Obsessive-compulsive disorder characterized by unwanted thoughts and repetitive behaviors"],
        ["P009", "Burnout", "State of emotional, physical, and mental exhaustion caused by excessive and prolonged stress"],
        ["P010", "Grief", "Natural response to loss, particularly the death of a loved one"],
        ["P011", "Low Self-Esteem", "Negative perception of oneself and feelings of inadequacy or unworthiness"],
        ["P012", "Anger Management", "Difficulty controlling anger responses in appropriate ways"],
        ["P013", "Relationship Issues", "Problems in interpersonal relationships causing distress"],
        ["P014", "Eating Concerns", "Unhealthy relationship with food, body image, or eating behaviors"],
        ["P015", "Substance Use", "Problematic use of alcohol, drugs, or other substances affecting wellbeing"]
    ]
    
    self_assessment_data = [
        ["question_id", "problem_id", "question_text", "response_type", "next_step"],
        ["Q001", "P001", "How often do you feel nervous or on edge?", "scale_1_5", "Q002"],
        ["Q002", "P001", "Do you find it difficult to control worrying?", "yes_no", "Q003"],
        ["Q003", "P001", "How much does anxiety interfere with your daily activities?", "scale_1_5", None],
        ["Q004", "P002", "How often do you feel down or hopeless?", "scale_1_5", "Q005"],
        ["Q005", "P002", "Have you lost interest or pleasure in activities you used to enjoy?", "yes_no", "Q006"],
        ["Q006", "P002", "How much does depression interfere with your daily activities?", "scale_1_5", None],
        ["Q007", "P003", "How often do you feel overwhelmed by responsibilities?", "scale_1_5", "Q008"],
        ["Q008", "P003", "Do you experience physical symptoms of stress (headaches, muscle tension)?", "yes_no", None],
        ["Q009", "P004", "How many hours of sleep do you typically get per night?", "numeric", "Q010"],
        ["Q010", "P004", "How often do you have trouble falling asleep?", "scale_1_5", None],
        ["Q011", "P005", "How often do you feel lonely?", "scale_1_5", "Q012"],
        ["Q012", "P005", "How many meaningful social interactions do you have per week?", "numeric", None],
        ["Q013", "P006", "How often do you experience sudden episodes of intense fear?", "scale_1_5", "Q014"],
        ["Q014", "P006", "During these episodes, do you experience physical symptoms like racing heart, sweating, or shortness of breath?", "yes_no", "Q015"],
        ["Q015", "P006", "Do you avoid certain situations due to fear of having a panic attack?", "yes_no", None],
        ["Q016", "P007", "Have you experienced or witnessed a traumatic event?", "yes_no", "Q017"],
        ["Q017", "P007", "Do you have recurring, unwanted memories of the traumatic event?", "scale_1_5", "Q018"],
        ["Q018", "P007", "Do you avoid things that remind you of the traumatic event?", "scale_1_5", None],
        ["Q019", "P008", "Do you experience unwanted, intrusive thoughts that cause anxiety?", "scale_1_5", "Q020"],
        ["Q020", "P008", "Do you feel compelled to perform certain behaviors or mental acts repeatedly?", "scale_1_5", "Q021"],
        ["Q021", "P008", "How much do these thoughts and behaviors interfere with your daily life?", "scale_1_5", None],
        ["Q022", "P009", "Do you feel emotionally drained from your work or responsibilities?", "scale_1_5", "Q023"],
        ["Q023", "P009", "Have you become more cynical or detached from your work or responsibilities?", "scale_1_5", "Q024"],
        ["Q024", "P009", "Do you feel less accomplished or effective in your work or responsibilities?", "scale_1_5", None],
        ["Q025", "P010", "Have you recently experienced a significant loss?", "yes_no", "Q026"],
        ["Q026", "P010", "How much does your grief interfere with your daily activities?", "scale_1_5", "Q027"],
        ["Q027", "P010", "Have you been able to find moments of peace or acceptance regarding your loss?", "scale_1_5", None],
        ["Q028", "P011", "How often do you criticize yourself?", "scale_1_5", "Q029"],
        ["Q029", "P011", "Do you often compare yourself unfavorably to others?", "scale_1_5", "Q030"],
        ["Q030", "P011", "Do you have difficulty accepting compliments or praise?", "scale_1_5", None],
        ["Q031", "P012", "How often do you feel angry or irritable?", "scale_1_5", "Q032"],
        ["Q032", "P012", "Do you have difficulty controlling your anger once it arises?", "scale_1_5", "Q033"],
        ["Q033", "P012", "Has your anger caused problems in your relationships or other areas of life?", "yes_no", None],
        ["Q034", "P013", "Are you experiencing conflict in an important relationship?", "yes_no", "Q035"],
        ["Q035", "P013", "How satisfied are you with your communication in this relationship?", "scale_1_5", "Q036"],
        ["Q036", "P013", "Do you feel understood and supported in this relationship?", "scale_1_5", None],
        ["Q037", "P014", "Are you preoccupied with thoughts about food, weight, or body shape?", "scale_1_5", "Q038"],
        ["Q038", "P014", "Do you engage in restrictive eating, binging, or purging behaviors?", "scale_1_5", "Q039"],
        ["Q039", "P014", "How much do concerns about food or body image interfere with your daily life?", "scale_1_5", None],
        ["Q040", "P015", "How often do you use alcohol or other substances?", "scale_1_5", "Q041"],
        ["Q041", "P015", "Have you tried to cut down on your substance use but found it difficult?", "yes_no", "Q042"],
        ["Q042", "P015", "Has your substance use caused problems in your relationships, work, or other areas of life?", "yes_no", None]
    ]
    
    suggestions_data = [
        ["suggestion_id", "problem_id", "suggestion_text", "resource_link"],
        ["S001", "P001", "Practice deep breathing exercises for 5 minutes when feeling anxious", "https://www.healthline.com/health/breathing-exercises-for-anxiety"],
        ["S002", "P001", "Try progressive muscle relaxation to reduce physical tension", "https://www.verywellmind.com/progressive-muscle-relaxation-pmr-2584097"],
        ["S003", "P001", "Consider speaking with a mental health professional about anxiety management strategies", None],
        ["S004", "P002", "Establish a daily routine with small, achievable goals", None],
        ["S005", "P002", "Engage in physical activity for at least 30 minutes daily", "https://www.mayoclinic.org/diseases-conditions/depression/in-depth/depression-and-exercise/art-20046495"],
        ["S006", "P002", "Consider speaking with a mental health professional about depression treatment options", None],
        ["S007", "P003", "Practice mindfulness meditation for 10 minutes daily", "https://www.mindful.org/how-to-practice-mindfulness/"],
        ["S008", "P003", "Identify and limit stress triggers in your daily life", None],
        ["S009", "P004", "Establish a consistent sleep schedule, even on weekends", None],
        ["S010", "P004", "Create a relaxing bedtime routine without screens", "https://www.sleepfoundation.org/sleep-hygiene/bedtime-routine-for-adults"],
        ["S011", "P005", "Join a community group or class based on your interests", None],
        ["S012", "P005", "Schedule regular video calls with friends or family members", None],
        ["S013", "P006", "Learn to recognize the early signs of a panic attack", "https://www.mind.org.uk/information-support/types-of-mental-health-problems/panic-attacks/"],
        ["S014", "P006", "Practice grounding techniques like the 5-4-3-2-1 method during panic attacks", "https://www.healthline.com/health/grounding-techniques"],
        ["S015", "P006", "Consider cognitive-behavioral therapy (CBT) which is highly effective for panic disorder", None],
        ["S016", "P007", "Practice trauma-focused mindfulness with professional guidance", None],
        ["S017", "P007", "Try journaling about your feelings related to the traumatic event", "https://psychcentral.com/health/benefits-of-journaling-for-ptsd"],
        ["S018", "P007", "Consider EMDR (Eye Movement Desensitization and Reprocessing) therapy with a qualified professional", "https://www.emdr.com/what-is-emdr/"],
        ["S019", "P008", "Learn to recognize and label intrusive thoughts without judgment", None],
        ["S020", "P008", "Practice exposure and response prevention techniques with professional guidance", "https://iocdf.org/about-ocd/treatment/erp/"],
        ["S021", "P008", "Consider joining an OCD support group to connect with others who understand", None],
        ["S022", "P009", "Set clear boundaries between work and personal life", None],
        ["S023", "P009", "Schedule regular breaks and time for activities you enjoy", None],
        ["S024", "P009", "Consider whether changes to your work environment or responsibilities might be beneficial", "https://www.helpguide.org/articles/stress/burnout-prevention-and-recovery.htm"],
        ["S025", "P010", "Allow yourself to feel grief without judgment", None],
        ["S026", "P010", "Create rituals or memorials to honor your loss", None],
        ["S027", "P010", "Consider joining a grief support group", "https://grief.com/grief-support-groups/"],
        ["S028", "P011", "Practice positive self-talk and challenge negative self-perceptions", None],
        ["S029", "P011", "Keep a journal of your accomplishments and positive qualities", None],
        ["S030", "P011", "Set small, achievable goals to build confidence through success", "https://www.mind.org.uk/information-support/types-of-mental-health-problems/self-esteem/"],
        ["S031", "P012", "Practice identifying your anger triggers", None],
        ["S032", "P012", "Learn and use time-out strategies when you feel anger escalating", "https://www.apa.org/topics/anger/control"],
        ["S033", "P012", "Consider anger management classes or therapy", None],
        ["S034", "P013", "Practice active listening techniques in your conversations", "https://www.verywellmind.com/what-is-active-listening-3024343"],
        ["S035", "P013", "Use 'I' statements to express feelings without blame", None],
        ["S036", "P013", "Consider couples or family therapy to improve communication", None],
        ["S037", "P014", "Practice mindful eating by paying attention to hunger cues and eating experience", "https://www.healthline.com/nutrition/mindful-eating-guide"],
        ["S038", "P014", "Challenge negative thoughts about body image", None],
        ["S039", "P014", "Consider working with a registered dietitian and therapist specialized in eating concerns", "https://www.nationaleatingdisorders.org/"],
        ["S040", "P015", "Keep a journal of your substance use patterns and triggers", None],
        ["S041", "P015", "Explore healthy alternatives to cope with stress or difficult emotions", None],
        ["S042", "P015", "Consider speaking with a healthcare provider about support options for substance use", "https://www.samhsa.gov/find-help/national-helpline"]
    ]
    
    feedback_prompts_data = [
        ["prompt_id", "stage", "prompt_text", "next_action"],
        ["FP001", "initial", "How are you feeling today?", "NA001"],
        ["FP002", "assessment", "Based on your responses, it seems you might be experiencing anxiety. Would you like to learn some coping strategies?", "NA002"],
        ["FP003", "assessment", "Based on your responses, it seems you might be experiencing depression. Would you like to learn some coping strategies?", "NA002"],
        ["FP004", "assessment", "Based on your responses, it seems you might be experiencing high stress. Would you like to learn some stress management techniques?", "NA002"],
        ["FP005", "assessment", "Based on your responses, it seems you might be experiencing sleep issues. Would you like to learn some sleep improvement strategies?", "NA002"],
        ["FP006", "assessment", "Based on your responses, it seems you might be experiencing social isolation. Would you like to learn some ways to increase social connection?", "NA002"],
        ["FP007", "follow_up", "How have the suggested strategies been working for you?", "NA003"],
        ["FP008", "follow_up", "Would you like to try different strategies for managing your symptoms?", "NA004"],
        ["FP009", "follow_up", "Would you like to speak with a mental health professional for additional support?", "NA005"],
        ["FP010", "assessment", "Based on your responses, it seems you might be experiencing panic attacks. Would you like to learn some techniques to manage them?", "NA002"],
        ["FP011", "assessment", "Based on your responses, it seems you might be experiencing symptoms of PTSD. Would you like to learn about coping strategies?", "NA002"],
        ["FP012", "assessment", "Based on your responses, it seems you might be experiencing OCD symptoms. Would you like to learn about management techniques?", "NA002"],
        ["FP013", "assessment", "Based on your responses, it seems you might be experiencing burnout. Would you like to learn about recovery strategies?", "NA002"],
        ["FP014", "assessment", "Based on your responses, it seems you might be experiencing grief. Would you like to learn about coping with loss?", "NA002"],
        ["FP015", "assessment", "Based on your responses, it seems you might be experiencing low self-esteem. Would you like to learn about building confidence?", "NA002"],
        ["FP016", "assessment", "Based on your responses, it seems you might be experiencing anger management issues. Would you like to learn about controlling anger?", "NA002"],
        ["FP017", "assessment", "Based on your responses, it seems you might be experiencing relationship issues. Would you like to learn about improving communication?", "NA002"],
        ["FP018", "assessment", "Based on your responses, it seems you might be experiencing concerns related to eating or body image. Would you like to learn about developing a healthier relationship with food and your body?", "NA002"],
        ["FP019", "assessment", "Based on your responses, it seems you might be experiencing concerns related to substance use. Would you like to learn about support options?", "NA002"],
        ["FP020", "follow_up", "Have you been able to practice the suggested techniques regularly?", "NA003"],
        ["FP021", "follow_up", "What challenges have you faced in implementing the suggested strategies?", "NA004"],
        ["FP022", "follow_up", "Would you like to try a different approach to managing your symptoms?", "NA004"],
        ["FP023", "follow_up", "Would you like to learn about additional resources for support?", "NA005"],
        ["FP024", "initial", "What specific mental health concerns would you like to address today?", "NA001"]
    ]
    
    next_actions_data = [
        ["action_id", "label", "description"],
        ["NA001", "Begin Assessment", "Start the self-assessment process to identify potential issues"],
        ["NA002", "View Suggestions", "Show personalized suggestions based on assessment results"],
        ["NA003", "Continue Current Plan", "Continue with the current strategies and check in later"],
        ["NA004", "Try New Strategies", "Provide alternative coping strategies and resources"],
        ["NA005", "Connect to Professional", "Provide information on accessing professional mental health support"],
        ["NA006", "Explore Resources", "Provide educational resources about specific mental health topics"],
        ["NA007", "Schedule Follow-up", "Set a time for a follow-up conversation to check progress"],
        ["NA008", "Crisis Support", "Provide immediate resources for Crisis situations"],
        ["NA009", "Self-care Plan", "Help develop a personalized self-care routine"],
        ["NA010", "Community Support", "Provide information about support groups and community resources"]
    ]
    
    finetuning_examples_data = [
        ["id", "prompt", "completion", "problem", "conversation_id"],
        ["FT001", "I've been feeling really anxious lately and can't seem to relax.", "It sounds like you're experiencing anxiety. Would you like to take a brief assessment to help me understand your symptoms better?", "P001", "C001"],
        ["FT002", "Yes, I'd like to take the assessment.", "Great. Let's start with a few questions about your anxiety symptoms. How often do you feel nervous or on edge?", "P001", "C001"],
        ["FT003", "I feel nervous almost every day.", "That's quite frequent. Do you find it difficult to control worrying?", "P001", "C001"],
        ["FT004", "Yes, I can't stop worrying about everything.", "I understand that must be challenging. How much does your anxiety interfere with your daily activities?", "P001", "C001"],
        ["FT005", "It's making it hard to focus at work and sleep at night.", "Based on your responses, it seems you're experiencing significant anxiety symptoms. Would you like to learn some coping strategies that might help?", "P001", "C001"],
        ["FT006", "I've been feeling really sad and don't enjoy anything anymore.", "I'm sorry to hear you're feeling this way. It sounds like you might be experiencing symptoms of depression. Would you like to take a brief assessment to help me understand better?", "P002", "C002"],
        ["FT007", "I'm having trouble sleeping and feel exhausted all the time.", "Sleep issues can significantly impact your well-being. Would you like to take a brief assessment about your sleep patterns to help me provide better suggestions?", "P004", "C003"],
        ["FT008", "I feel like I have no one to talk to and I'm always alone.", "It sounds like you might be experiencing social isolation. Would you like to discuss some ways to increase meaningful social connections in your life?", "P005", "C004"],
        ["FT009", "I'm so stressed with work and family responsibilities.", "Managing multiple responsibilities can certainly be stressful. Would you like to take a brief assessment about your stress levels to help me provide targeted suggestions?", "P003", "C005"],
        ["FT010", "The deep breathing exercises helped a lot with my anxiety.", "I'm glad to hear the deep breathing exercises have been helpful! Would you like to learn additional techniques to manage anxiety?", "P001", "C001"],
        ["FT011", "I keep having these moments where my heart races and I can't breathe.", "What you're describing sounds like it could be panic attacks. Would you like to take a brief assessment to help me understand your experiences better?", "P006", "C006"],
        ["FT012", "I was in a car accident last month and keep having nightmares about it.", "I'm sorry to hear about your accident. What you're describing could be related to trauma. Would you like to take a brief assessment to help me understand what you're experiencing?", "P007", "C007"],
        ["FT013", "I have to check that my door is locked exactly three times before I can leave.", "That sounds like it might be related to OCD. Would you like to take a brief assessment to help me understand your experiences better?", "P008", "C008"],
        ["FT014", "I feel completely exhausted by my job and don't care about it anymore.", "What you're describing sounds like it could be burnout. Would you like to take a brief assessment to help me understand your situation better?", "P009", "C009"],
        ["FT015", "My mother died recently and I'm having a hard time coping.", "I'm very sorry for your loss. Grief is a natural response to losing someone important to you. Would you like to discuss some ways to cope with grief?", "P010", "C010"],
        ["FT016", "I always feel like I'm not good enough compared to everyone else.", "It sounds like you might be experiencing low self-esteem. Would you like to take a brief assessment to help me understand your feelings better?", "P011", "C011"],
        ["FT017", "I get angry so easily and sometimes I say things I regret.", "Managing anger can be challenging. Would you like to take a brief assessment to help me understand your anger patterns better?", "P012", "C012"],
        ["FT018", "My partner and I keep having the same argument over and over.", "Recurring conflicts can be frustrating. Would you like to discuss some communication strategies that might help improve your relationship?", "P013", "C013"],
        ["FT019", "I'm constantly thinking about my weight and feel guilty after eating.", "Your relationship with food and body image sounds like it might be causing you distress. Would you like to take a brief assessment to help me understand better?", "P014", "C014"],
        ["FT020", "I've been drinking more than usual to cope with stress.", "Using substances to cope with stress can sometimes lead to other issues. Would you like to take a brief assessment about your substance use patterns?", "P015", "C015"],
        ["FT021", "The mindfulness techniques have really helped with my stress levels.", "That's wonderful to hear! Mindfulness can be very effective for managing stress. Would you like to learn some additional stress management techniques?", "P003", "C005"],
        ["FT022", "I tried the sleep routine you suggested but I'm still having trouble.", "I'm sorry to hear the routine hasn't fully helped yet. Sleep issues can sometimes require trying different approaches. Would you like to discuss some alternative strategies?", "P004", "C003"],
        ["FT023", "I joined a local book club like you suggested and it's really helping with my loneliness.", "That's wonderful to hear! Social connections through shared interests can be very beneficial. How has the experience been for you so far?", "P005", "C004"],
        ["FT024", "I'm having a really hard time right now and don't know what to do.", "I'm sorry you're going through a difficult time. Can you tell me a bit more about what you're experiencing so I can better understand how to support you?", None, "C016"],
        ["FT025", "Sometimes I feel so overwhelmed I just want to give up.", "I'm really concerned about what you're sharing. Your feelings are valid, and it's important you know that help is available. Would you like to talk about Crisis support resources?", None, "C017"]
    ]
    
    # Create sheets and add data
    sheets_data = {
        "Problems": problems_data,
        "SelfAssessment": self_assessment_data,
        "Suggestions": suggestions_data,
        "FeedbackPrompts": feedback_prompts_data,
        "NextActions": next_actions_data,
        "FinetuningExamples": finetuning_examples_data
    }
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create and populate sheets
    for sheet_name, data in sheets_data.items():
        ws = wb.create_sheet(sheet_name)
        
        # Add data to sheet
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply header styling
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Apply border to all cells
                cell.border = border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "expanded_mental_health_kb_data.xlsx")
    wb.save(output_path)
    print(f"Excel file created at: {output_path}")
    return output_path


if __name__ == "__main__":
    create_expanded_mental_health_kb_excel()